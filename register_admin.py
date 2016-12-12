from flask import Flask
import flask
from flask import render_template
from flask import request,redirect,url_for
#from werkzeug import secure_filename
import openpyxl
from db2 import *
import os
from flask import flash
from datetime import *
from flask.ext.login import LoginManager,login_required,current_user,logout_user
import flask.ext.login as flask_login
from flask import send_from_directory
UPLOAD_FOLDER=os.getcwd()

app=Flask(__name__)
app.config['SECRET_KEY']='hard to guess'
app.config['UPLOAD_FOLDER']=UPLOAD_FOLDER

login_manager.init_app(app)

class User(flask_login.UserMixin):
    pass


@login_manager.user_loader 
def user_loader(email2):
    print '1'
    try:
        if(registration.objects.get(email_db=email2)):
            user=User()
            user.id=email2
            return user
    except DoesNotExist:
        return
    if not registration.objects(email_db=email2):
        return

@app.route('/')
def first():
    return render_template('first.html')

@app.route('/register',methods=['GET','POST'])
def register():
    if request.method=='GET':
        return render_template('register.html')
    else:
        
        name=request.form['Name']
        email=request.form['Email']
        password=request.form['Password']
        confirm=request.form['Confirm_Password']
        try:
            if (registration.objects.get(email_db=email) ):
                return redirect(url_for('register'))
        except DoesNotExist:
            if (password==confirm):
                try:
                    reg=registration(name_db=name,email_db=email,create_password=password).save()
                    flash('registration complete pls do the login')
                    return redirect(url_for('login'))
                except Exception,e:
                    return redirect(url_for('register'))
            else:
                flash("password mismatch")
                return redirect(url_for('register'))
        except Exception,e:
            return redirect(url_for('register'))
        
            


                
@app.route('/login',methods=['GET','POST'])
def login():
   
    if request.method=='GET':
        return render_template('login.html')
    else:
        email2=request.form['EMAILID']

        password2=request.form['PASSWORD']
   

        try:
            if (registration.objects.get(email_db=email2) ):
                try:
                    if (registration.objects.get(email_db=email2,create_password=password2) ):


                
                        user = User()
                        user.id = email2
                        print '3'
                        flask_login.login_user(user)                                 
                        return redirect(url_for('bookmark1'))
                except DoesNotExist:

                    flash("email password combination mismatch please try again")
                    return redirect(url_for('login'))
        except DoesNotExist:
            flash("email not registered please do registration ")
            return redirect(url_for('first'))
                
        if not registration.objects(email_db=email2):
            flash("email not registered please do the registration")

            return redirect(url_for('first'))

@app.route('/bookmark1')
def bookmark1():
    
    return render_template('bookmark1.html')
    
@app.route('/add_bookmark',methods=['GET','POST'])
@flask_login.login_required
def add_bookmark():
    current_email= flask.ext.login.current_user.id
    if request.method=='GET':
        return render_template('bookmark_form.html')
    else:
        name=request.form['Name']
        url=request.form['URL']
        label=request.form['Label']
        notes=request.form['Notes']
        try:
            if Bookmark_db.objects.get(location=url,email=current_email):

                flash('Bookmark with same url is already present')
                return redirect(url_for('add_bookmark'))
        except DoesNotExist:
            current_email= flask.ext.login.current_user.id
            bookmark_db=Bookmark_db(name=name,location=url,labels=label,notes=notes,email=current_email).save()
            return redirect(url_for('views'))
        except Exception,e:
            flash(str(e))
            return redirect(url_for('add_bookmark'))
@app.route('/views')
@flask_login.login_required
def views():
    current_email= flask.ext.login.current_user.id
    list1=[]
    for i in Bookmark_db.objects(email=current_email):
        list1.append(i.location)
        print list1
    
    if not Bookmark_db.objects(email=current_email):
        flash('No Bookmark available')
        return render_template('bookmark_view.html',url_list=list1)
    return render_template('bookmark_view.html',url_list=list1)



@app.route('/delete')
@flask_login.login_required
def delete():
    current_email= flask.ext.login.current_user.id
    for i in Bookmark_db.objects(email=current_email):

        i.delete()
        flash("Bookmark deleted successfully")
    
    if not Bookmark_db.objects(email=current_email):
        

        return redirect(url_for('views'))
    return redirect(url_for('views'))

@app.route('/edit/<aa>',methods=["GET","POST"])
@flask_login.login_required
def edit(aa): 
    if request.method=="GET":
        current_email= flask.ext.login.current_user.id
        c=Bookmark_db.objects.get(email=current_email,location="http://"+aa)
        d= c.name
        f=c.labels
        g=c.notes

        return render_template('bookmark_form2.html',d=d,f=f,g=g)
    else:
        current_email= flask.ext.login.current_user.id
        name=request.form['Name']
        label=request.form['Label']
        notes=request.form['Notes']

        up=Bookmark_db.objects.get(email=current_email,location="http://"+aa)
        up.update(name=name,labels=label,notes=notes)
        flash("changes successfully applied")
        return redirect(url_for('views'))


@app.route('/export',methods=['GET','POST'])
@flask_login.login_required
def export():
    if request.method=='GET':
        current_email= flask.ext.login.current_user.id

        list_url=[]
        for i in Bookmark_db.objects(email=current_email):
            list_url.append(i.location)
        
        if not Bookmark_db.objects(email=current_email):
            flash('No Bookmark available')
            return render_template('bookmark_view.html')
        return render_template('bookmark_view2.html',url_list=list_url)        
    else:
        url_list=request.form.getlist('url')
        print url_list
        try:
            wb=openpyxl.Workbook()
            sheet=wb.get_active_sheet()
            print sheet
        except Exception,e:
            print str(e)
        current_email=  flask.ext.login.current_user.id
        print "log check 1"
        for j in url_list:

            for i in Bookmark_db.objects(email=current_email,location=j):

                li=[]
                li=i._data.keys()

                break
        if not url_list:
            flash ('please select atleast one option')
            return redirect(url_for('export'))
        for k in range(1,len(li)):
            sheet.cell(row=1,column=k).value=li[k-1]
            print "log check 2"
        row_num=2
        for i in url_list:
            print i

            for j in Bookmark_db.objects(email=current_email,location=i):
                li1=[]
            
                li1=j._data.values()
                
                for l in range(1,len(li1)):
                    sheet.cell(row=row_num,column=l).value=li1[l-1]
                row_num+=1
        try:

            wb.save(UPLOAD_FOLDER+'/'+current_email+'.xlsx')
        except Exception,e:

        flash('files saved successfully')
        return render_template('download.html',i=current_email+'.xlsx')


@app.route('/uploaded/<filename>')
#@flask_login.login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'],filename)

@app.route('/import_file' , methods=['GET','POST'])
@flask_login.login_required
def import_file():
    if request.method=='GET':
        return render_template('upload.html')
    else:
        current_email=flask.ext.login.current_user.id
        file1=request.files['file_name']
            

        wb = openpyxl.load_workbook('/home/sumit/google_bookmark/'+str(file1.filename))
        sheet=wb.get_active_sheet()
        d=sheet.get_highest_row()
        for i in range(2,d+1):
            try:
                if (Bookmark_db.objects.get(email=current_email,location=sheet.cell(row =i,column=5).value)):
                    continue
            except DoesNotExist:

                bookmark_db=Bookmark_db(name=sheet.cell(row =i,column=1).value,location=sheet.cell(row =i,column=5).value,
                    labels=sheet.cell(row =i,column=3).value,notes=sheet.cell(row =i,column=2).value,email=current_email).save()
        return redirect(url_for('views'))

@app.route('/remove/<aa>')
@flask_login.login_required
def remove(aa):
    
    current_email= flask.ext.login.current_user.id
    c=Bookmark_db.objects.get(email=current_email,location="http://"+aa)
    c.delete()    
    flash("File Delted Successfully")
    return redirect(url_for('views'))

@app.route('/logout')


def logout():
    flask_login.logout_user()
    flash("successfully Logged Out")
    return redirect (url_for('first'))


@login_manager.unauthorized_handler
def unauthorized_handler():

    flash('Unauthorized User Please Register Or login')
    return redirect (url_for('first'))


if __name__ == '__main__':
    app.run(debug=True)